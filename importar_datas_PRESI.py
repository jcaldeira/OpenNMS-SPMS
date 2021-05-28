"""
https://stackoverflow.com/questions/2600775/how-to-get-week-number-in-python
https://stackoverflow.com/questions/24370385/how-to-format-cell-with-datetime-object-of-the-form-yyyy-mm-dd-hhmmss-in-exc
https://stackoverflow.com/questions/28517508/read-excel-cell-value-and-not-the-formula-computing-it-openpyxl
"""

import datetime
import logging
import os
import subprocess
import sys
import time
from copy import copy

import openpyxl
from openpyxl.styles import NamedStyle
from openpyxl.styles import Alignment, Border, Side

main_logger = logging.getLogger(__name__)
main_logger.setLevel(logging.DEBUG)

file_formatter = logging.Formatter('%(asctime)s||%(levelname)s||%(name)s||%(message)s')
stream_formatter = logging.Formatter('%(message)s')

file_handler = logging.FileHandler(f'{os.path.join("Logs",os.path.splitext(os.path.basename(__file__))[0])}.log')
file_handler.setLevel(logging.DEBUG)
file_handler.setFormatter(file_formatter)

stream_handler = logging.StreamHandler()
stream_handler.setLevel(logging.INFO)
stream_handler.setFormatter(stream_formatter)

main_logger.addHandler(file_handler)
main_logger.addHandler(stream_handler)


def main():
    if sys.platform in ('linux', 'darwin'):
        excel = '/mnt/c/Users/joao.caldeira.ext/SPMS - Serviços Partilhados do Ministério da Saúde, EPE/SPMS DSI RDIS - RIS Corporate - RIS2020 Cadastro/RIS2020 - Cadastro.xlsx'
        excel2 = '/mnt/c/Users/joao.caldeira.ext/OneDrive - Portugal Telecom/Trabalho/COS/SPMS/PRESI/Calendarização-quattro.xlsx'
    elif sys.platform == 'win32':
        excel = r"C:\Users\joao.caldeira.ext\SPMS - Serviços Partilhados do Ministério da Saúde, EPE\SPMS DSI RDIS - RIS Corporate - RIS2020 Cadastro\RIS2020 - Cadastro.xlsx"
        # excel = r"C:\Users\joao.caldeira.ext\OneDrive - Portugal Telecom\Trabalho\COS\SPMS\PRESI\RIS2020 - Cadastro - Copy.xlsx"
        excel2 = r"C:\Users\joao.caldeira.ext\OneDrive - Portugal Telecom\Trabalho\COS\SPMS\PRESI\Calendarização-quattro.xlsx"

    excelFileName = os.path.basename(excel)
    pathToExcelFile = os.path.dirname(excel)

    excelFileName2 = os.path.basename(excel2)
    pathToExcelFile2 = os.path.dirname(excel2)

    # Open Excel
    try:
        wb_values = openpyxl.load_workbook(excel, data_only=True)
        wb = openpyxl.load_workbook(excel)
        delTempFileFlag = False
    except PermissionError as e:
        main_logger.info(f'File is in use, so cannot be accessed:\n{e.filename}\n')
        createTempFile(pathToExcelFile, os.getcwd(), excelFileName)
        main_logger.info('Temporary copy of the file created to work with\n')
        wb_values = openpyxl.load_workbook(excelFileName, data_only=True)
        wb = openpyxl.load_workbook(excelFileName)
        delTempFileFlag = True
    except FileNotFoundError as e:
        sys.exit(f'File not found: {e.filename}')


    # Open Excel 2
    try:
        wb2 = openpyxl.load_workbook(excel2, data_only=True)
        delTempFileFlag2 = False
    except PermissionError as e:
        main_logger.info(f'File is in use, so cannot be accessed:\n{e.filename}\n')
        createTempFile(pathToExcelFile2, os.getcwd(), excelFileName2)
        main_logger.info('Temporary copy of the file created to work with\n')
        wb2 = openpyxl.load_workbook(excelFileName2, data_only=True)
        delTempFileFlag2 = True
    except FileNotFoundError as e:
        sys.exit(f'File not found: {e.filename}')


    sheet_values = wb_values['RIS2020']
    sheet = wb['RIS2020']
    sheet2 = wb2[f"Week {datetime.date.today().strftime('%V')}"]

    style = {
        'number_format': 'dd-mmm',
        'alignment': Alignment(horizontal='center', vertical='center'),
        'border_style': Side(border_style='thin'),
    }
    style['border'] = Border(
        left=style['border_style'],
        right=style['border_style'],
        top=style['border_style'],
        bottom=style['border_style']
    )

    lista_sites = []

    # # Get all sites with schedule for this week
    for row in sheet2.iter_rows(min_row=2, values_only=True):
        if type(row[5]) == datetime.datetime and str(row[10]).upper() == 'CONFIRMADO':
            site_info = {
                'site_id': str(row[0]).zfill(4),
                'date_presi': row[5].date(),
            }

            lista_sites.append(site_info)

    for site in lista_sites:
        for row_index, row2 in enumerate(sheet_values.iter_rows(min_row=3, values_only=True), 3):
            if site['site_id'] == row2[0] and row2[26] == None:
                sheet[f'AA{row_index}'].value = site['date_presi']
                sheet[f'AA{row_index}'].alignment = style['alignment']
                sheet[f'AA{row_index}'].border = style['border']
                sheet[f'AA{row_index}'].number_format = style['number_format']


    try:
        wb.save(excel)
    except:
        main_logger.info('Was not possible to save the file\n')
        input('Try again? <Press ENTER>')
        wb.save(excel)

    wb_values.close()
    wb.close()
    wb2.close()
    if delTempFileFlag == True:
        os.remove(excelFileName)
    if delTempFileFlag2 == True:
        os.remove(excelFileName2)

    return len(lista_sites)



def cls():
    if sys.platform in ("linux", "darwin"):
        subprocess.run(["clear"])
    elif sys.platform == "win32":
        subprocess.run(["cls"], shell=True)



def createTempFile(pathToExcelFile, cwd, excelFileName):
    if sys.platform in ('linux', 'darwin'):
        subprocess.run(['cp', pathToExcelFile, cwd, excelFileName])
    elif sys.platform == 'win32':
        subprocess.run(['robocopy', pathToExcelFile, cwd, excelFileName], shell=True, stdout=subprocess.DEVNULL)

    # tempFile = f'{os.path.splitext(excelFileName)[0]}.temp{os.path.splitext(excelFileName)[1]}'
    # os.rename(excelFileName, f'temp-{tempFile}')




if __name__ == '__main__':
    cls()
    time_start = datetime.datetime.now()
    time_start = time_start.strftime('%H:%M:%S %d/%m/%Y')
    perf_counter_start = time.perf_counter()

    sites = main()

    perf_counter_stop = time.perf_counter()
    time_stop = datetime.datetime.now()
    time_stop = time_stop.strftime('%H:%M:%S %d/%m/%Y')

    main_logger.info(f'\nNumber of sites: {sites}')
    main_logger.info(f'Finished in {(perf_counter_stop - perf_counter_start):.3f} second(s)')
    main_logger.info(f'Started at {time_start} and finished at {time_stop}\n')
