"""
https://stackoverflow.com/questions/56770093/read-from-excel-file-that-is-open-in-python
"""

import datetime
import ipaddress
import os
import subprocess
import sys

import openpyxl


def main():
    excel = r"C:\Users\joao.caldeira.ext\SPMS - Serviços Partilhados do Ministério da Saúde, EPE\SPMS DSI RDIS - RIS Corporate - RIS2020 Cadastro\RIS2020 - Cadastro.xlsx"
    excelFileName = os.path.basename(excel)
    pathToExcelFile = os.path.dirname(excel)

    chars_a_remover = ['á','Á','à','À','ã','Ã','â','Â','é','É','è','È','ê','Ê','í','Í','ó','Ó','õ','Õ','ô','Ô','ú','Ú','û','Û','ù','Ù','ç','s/n']
    chars_a_inserir = ['a','A','a','A','a','A','a','A','e','E','e','E','e','E','i','I','o','O','o','O','o','O','u','U','u','U','u','U','c','']

    try:
        wb = openpyxl.load_workbook(excel, data_only=True, read_only=True)
        delTempFileFlag = False
    except PermissionError as e:
        print(f"File is open, so cannot be accessed:\n{e.filename}\n")
        createTempFile(pathToExcelFile, os.getcwd(), excelFileName)
        print("Temporary copy of file created to work with\n")
        wb = openpyxl.load_workbook(excelFileName, data_only=True, read_only=True)
        delTempFileFlag = True

    except FileNotFoundError as e:
        sys.exit(f"File not found: {e.filename}")


    sheet = wb["RIS2020"]
    counter = 0
    excluded_ids = []
    excluded_ips = []
    for row in sheet.iter_rows(min_row=3, values_only=True):
        if type(row[26]) == datetime.datetime:
        # if row[26] is not None:
            data_presi = row[26].date()
            ip = str(row[24])
            hostname = str(row[23])
            site_id = hostname[:4]
            # if validate_ip(ip) and validate_date_inter(data_presi) and site_id not in excluded_ids and ip not in excluded_ips:
            # if validate_ip(ip) and hostname == '0073-B01-SW01':
            if validate_ip(ip) and site_id == '0287':
                if str(row[0]) == site_id:
                    entidade, morada, cp, localidade, latitude, longitude, requisition, modelo, serial, node_id = site_info(chars_a_remover, chars_a_inserir, row)

                    print(f"/opt/opennms/bin/provision.pl service add '{requisition}' {node_id} {ip} ICMP")
                    print(f"/opt/opennms/bin/provision.pl service add '{requisition}' {node_id} {ip} SNMP")
                    print(f"/opt/opennms/bin/provision.pl category remove '{requisition}' {node_id} '{entidade}'")
                    print(f"/opt/opennms/bin/provision.pl category add '{requisition}' {node_id} '{entidade.upper()}'")
                    print(f"/opt/opennms/bin/provision.pl category add '{requisition}' {node_id} 'Production'")
                    print(f"/opt/opennms/bin/provision.pl category add '{requisition}' {node_id} 'Switches'")
                    print(f"/opt/opennms/bin/provision.pl asset add '{requisition}' {node_id} address1 '{morada}'")
                    print(f"/opt/opennms/bin/provision.pl asset add '{requisition}' {node_id} zip '{cp}'")
                    print(f"/opt/opennms/bin/provision.pl asset add '{requisition}' {node_id} city '{localidade}'")
                    print(f"/opt/opennms/bin/provision.pl asset add '{requisition}' {node_id} modelNumber '{modelo}'")
                    print(f"/opt/opennms/bin/provision.pl asset add '{requisition}' {node_id} serialNumber '{serial}'")
                    print(f"/opt/opennms/bin/provision.pl asset add '{requisition}' {node_id} latitude '{latitude}'")
                    print(f"/opt/opennms/bin/provision.pl asset add '{requisition}' {node_id} longitude '{longitude}'")
                    print(f"/opt/opennms/bin/provision.pl asset add '{requisition}' {node_id} country Portugal")
                    print(f"/opt/opennms/bin/provision.pl requisition import {requisition}\n\n")

                    counter += 1

                else:
                    print(f"IDs do not match {site_id} (on switch name)\n")

    wb.close()
    if delTempFileFlag == True:
        os.remove(excelFileName)

    print(f'Generated {counter} configurations')


def site_info(chars_a_remover, chars_a_inserir, row):
    entidade = str(row[1]).replace(" ","_").strip()
    morada = str(row[4]).strip()
    cp = str(row[5]).strip()
    localidade = str(row[6]).strip()
    latitude = str(row[7]).strip()
    longitude = str(row[8]).strip()
    requisition = str(row[28]).strip()
    modelo = str(row[20]).strip()
    serial = str(row[21]).strip()
    node_id = str(row[27]).strip()
    for i in range(0,len(chars_a_remover)):
        morada = morada.lower().replace(chars_a_remover[i],chars_a_inserir[i]).strip().capitalize()
        localidade = localidade.lower().replace(chars_a_remover[i], chars_a_inserir[i]).strip().capitalize()
    return entidade, morada, cp, localidade, latitude, longitude, requisition, modelo, serial, node_id



def validate_date_inter(date):
    low_date_lim = datetime.date(2021, 5, 4)
    high_date_lim = datetime.date(2021, 5, 5)
    return low_date_lim <= date <= high_date_lim



def validate_ip(ip):
    try:
        ipaddress.ip_address(ip)
    except:
        return False
    else:
        return True



def cls():
    if sys.platform in ("linux", "darwin"):
        subprocess.run(["clear"])
    elif sys.platform == "win32":
        subprocess.run(["cls"], shell=True)



def createTempFile(pathToExcelFile, cwd, excelFileName):
    if sys.platform in ("linux", "darwin"):
        sys.exit(f"Copy mechanism of in use files not implemented in non Windows platforms")
    elif sys.platform == "win32":
        subprocess.run(["robocopy", pathToExcelFile, cwd, excelFileName], shell=True, stdout=subprocess.DEVNULL)



if __name__ == "__main__":
    cls()
    main()
