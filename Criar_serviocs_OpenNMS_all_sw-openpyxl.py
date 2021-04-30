"""Script criado por Rui Pereira e adaptado por João Caldeira"""

import os
import re
import os.path
import socket
import openpyxl
from os import path


if socket.gethostname() == 'LP013150':
    excel = 'C:\\Users\\rui.pereira.ext\\OneDrive - SPMS - Serviços Partilhados do Ministério da Saúde, EPE\\RIS2020 Cadastro\\RIS2020 - Cadastro.xlsx'
else:
    excel = 'D:\\SPMS OneDrive\\SPMS - Serviços Partilhados do Ministério da Saúde, EPE\\SPMS DSI RDIS - RIS Corporate - RIS2020 Cadastro\\RIS2020 - Cadastro.xlsx'

if not os.path.exists(excel):
    print ('Sem acesso ao ficheiro excel!')

# To open Workbook
wb = openpyxl.load_workbook(excel, data_only=True)
sheet = wb["RIS2020"]

sites_com_dois_routers = []
for row in sheet.iter_rows(min_row=3, values_only=True):
    hostname = str(row[23])
    ip = str(row[24])
    aces = str(row[28])
    site_id = hostname[0:4]
    chars_a_remover = ['á','à','ã','â','é','è','ê','í','ó','õ','ô','ú','û','ù','ç','s/n']
    chars_a_inserir = ['a','a','a','a','e','e','e','i','o','o','o','u','u','u','c','']
    if ip != '' and hostname == '2014-B01-SW01':
        if 'B01-SW01' not in hostname:
            for row2 in sheet.iter_rows(min_row=3, values_only=True):
                if str(row2[0]) == site_id:
                    entidade = str(row[1])
                    entidade = entidade.replace(" ","_")
                    entidade_upper = entidade.upper()
                    morada = str(row2[4])
                    cp = str(row2[5])
                    localidade = str(row2[6])
                    latitude = str(row2[7])
                    longitude = str(row2[8])
                    local = str(row2[3])
                    for i in range(0,len(chars_a_remover)):
                        morada = morada.replace(chars_a_remover[i],chars_a_inserir[i])
                        localidade = localidade.replace(chars_a_remover[i], chars_a_inserir[i])

        else:
            entidade = str(row[1])
            entidade = entidade.replace(" ","_")
            entidade_upper = entidade.upper()
            morada = str(row[4])
            cp = str(row[5])
            localidade = str(row[6])
            local = str(row[3])
            latitude = str(row[7])
            longitude = str(row[8])
            for i in range(0,len(chars_a_remover)):
                morada = morada.replace(chars_a_remover[i],chars_a_inserir[i])
                localidade = localidade.replace(chars_a_remover[i], chars_a_inserir[i])

        aces = str(row[28])
        modelo = str(row[20])
        serial = str(row[21])
        node_id = str(row[27])

        print ("/opt/opennms/bin/provision.pl service add '" + aces + "' " + node_id + " " + ip + " ICMP")
        print ("/opt/opennms/bin/provision.pl service add '" + aces + "' " + node_id + " " + ip + " SNMP")
        print ("/opt/opennms/bin/provision.pl category remove '" + aces + "' " + node_id + " '" + entidade + "'")
        print("/opt/opennms/bin/provision.pl category add '" + aces + "' " + node_id + " '" + str(entidade_upper) + "'")
        print ("/opt/opennms/bin/provision.pl category add '" + aces + "' " + node_id + " Production")
        print ("/opt/opennms/bin/provision.pl category add '" + aces + "' " + node_id + " Switches")
        print("/opt/opennms/bin/provision.pl asset add '" + aces + "' " + node_id + " address1 '" + morada + "'")
        print("/opt/opennms/bin/provision.pl asset add '" + aces + "' " + node_id + " zip " + cp)
        print("/opt/opennms/bin/provision.pl asset add '" + aces + "' " + node_id + " city '" + localidade + "'")
        print("/opt/opennms/bin/provision.pl asset add '" + aces + "' " + node_id + " modelNumber " + modelo)
        print("/opt/opennms/bin/provision.pl asset add '" + aces + "' " + node_id + " serialNumber " + serial)
        print ("/opt/opennms/bin/provision.pl asset add '" + aces + "' " + node_id + " latitude " + latitude)
        print ("/opt/opennms/bin/provision.pl asset add '" + aces + "' " + node_id + " longitude " + longitude)
        print("/opt/opennms/bin/provision.pl asset add '" + aces + "' " + node_id + " country Portugal")
        print ("/opt/opennms/bin/provision.pl requisition import '" + aces + "'\n")
