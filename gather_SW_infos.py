"""
Tem de ser executado no WLS
"""

import concurrent.futures
import datetime
import ipaddress
import logging
import os
import re
import subprocess
import sys
import time

import netmiko
import openpyxl
from netmiko.ssh_exception import (AuthenticationException,
                                   NetMikoTimeoutException)
from paramiko.ssh_exception import SSHException

# ▬▬▬▬▬▬▬▬▬▬▬▬▬▬▬ Logging ▬▬▬▬▬▬▬▬▬▬▬▬▬▬▬
main_logger = logging.getLogger(__name__) # criar um logger específico deste módulo
main_logger.setLevel(logging.DEBUG) # definir o nível de verbosidade do logger

file_formatter = logging.Formatter('%(asctime)s:%(levelname)s:%(name)s:%(message)s') # criar o formato dos logs a aplicar no file_formatter
steam_formatter = logging.Formatter('%(message)s') # criar o formato dos logs a aplicar no stream_handler

file_handler = logging.FileHandler(f'{os.path.join("Logs",os.path.splitext(os.path.basename(__file__))[0])}.log') # criar ficheiro de logs
file_handler.setLevel(logging.DEBUG) # definir o nível de verbosidade do file_handler
file_handler.setFormatter(file_formatter) # aplicar o formato de log anteriormente criado

stream_handler = logging.StreamHandler() # criar stream_handler que serve para mostrar logs noutros canais
stream_handler.setLevel(logging.INFO) # definir o nível de verbosidade do stream_handler
stream_handler.setFormatter(steam_formatter) # aplicar o formato de log anteriormente criado

main_logger.addHandler(file_handler) # aplicar file_handler ao logger
main_logger.addHandler(stream_handler) # aplicar stream_handler ao logger

# logging.basicConfig(filename='netmiko.log', level=logging.DEBUG)
# logger = logging.getLogger("netmiko")
# ▬▬▬▬▬▬▬▬▬▬▬▬▬▬▬ End Logging ▬▬▬▬▬▬▬▬▬▬▬▬▬▬▬


def validate_ip(ip):
    try:
        ipaddress.ip_address(ip)
    except:
        return False
    else:
        return True



def validate_date_inter(date):
    low_date_lim = datetime.date(2021, 5, 10)
    high_date_lim = datetime.date(2021, 5, 12)
    return low_date_lim <= date <= high_date_lim



def cls():
    if sys.platform in ("linux", "darwin"):
        subprocess.run(["clear"])
    elif sys.platform == "win32":
        subprocess.run(["cls"], shell=True)



def createTempFile(pathToExcelFile, cwd, excelFileName):
    if sys.platform in ("linux", "darwin"):
        subprocess.run(["cp", pathToExcelFile, cwd, excelFileName])
    elif sys.platform == "win32":
        subprocess.run(["robocopy", pathToExcelFile, cwd, f'{os.path.splitext(excelFileName)[0]}.temp{os.path.splitext(excelFileName)[1]}'], shell=True, stdout=subprocess.DEVNULL)



def env_exec():
    if sys.platform in ("linux", "darwin"):
        excel = "/mnt/c/Users/joao.caldeira.ext/SPMS - Serviços Partilhados do Ministério da Saúde, EPE/SPMS DSI RDIS - RIS Corporate - RIS2020 Cadastro/RIS2020 - Cadastro.xlsx"
    elif sys.platform == "win32":
        excel = r"C:\Users\joao.caldeira.ext\SPMS - Serviços Partilhados do Ministério da Saúde, EPE\SPMS DSI RDIS - RIS Corporate - RIS2020 Cadastro\RIS2020 - Cadastro.xlsx"

    excelFileName = os.path.basename(excel)
    pathToExcelFile = os.path.dirname(excel)

    username = 'jcaldeira'
    password = os.getenv('PWD_TACACS_RIS_SPMS')
    # password = os.getenv('PWD_TACACS_RIS_ALTICE')

    try:
        wb = openpyxl.load_workbook(excel, data_only=True, read_only=True)
        delTempFileFlag = False
    except PermissionError as e:
        print(f"File is open, so cannot be accessed:\n{e.filename}\n")
        createTempFile(pathToExcelFile, os.getcwd(), excelFileName)
        print("Temporary copy of file created to work with\n")
        wb = openpyxl.load_workbook(excelFileName, data_only=True, read_only=True)
        delTempFileFlag = True
    except FileNotFoundError as e:
        sys.exit(f"File not found: {e.filename}")

    sheet = wb["RIS2020"]
    device_list = []
    excluded_ids = []
    included_ids = []
    excluded_ips = []
    included_ips = []

    for row in sheet.iter_rows(min_row=3, values_only=True):
        if type(row[26]) == datetime.datetime:
        # if row[26] is not None:
            data_presi = row[26].date()
            ip = str(row[24])
            hostname = str(row[23])
            site_id = row[0]
            # if validate_ip(ip) and validate_date_inter(data_presi) and site_id not in excluded_ids and ip not in excluded_ips:
            # if validate_ip(ip) and hostname == '0073-B01-SW01':
            if validate_ip(ip) and site_id == '0345':
            # if validate_ip(ip) and site_id in included_ids or ip in included_ips:
                # if str(hostname[:4]) == site_id:
                    equipment = {
                        'device_type': 'cisco_ios',
                        'ip': ip,
                        'username': username,
                        'password': password,
                        'secret': site_id,
                        'ssh_config_file': '~/.ssh/config'
                    }

                    device_list.append(equipment)


    main_logger.debug(f'device_list: {device_list}')


    with concurrent.futures.ThreadPoolExecutor() as executor:
        executor.map(connect_and_commands, device_list)

    wb.close()
    if delTempFileFlag == True:
        os.remove(excelFileName)

    return len(device_list)



def connect_and_commands(equipment):
    main_logger.info(f"Accessing: {equipment['secret']} ({equipment['ip']})")
    try:
        with netmiko.ConnectHandler(**equipment) as connection:
            command_string = 'show version'
            output = connection.send_command(command_string = command_string)
            hostname = connection.find_prompt()

    except NetMikoTimeoutException:
        main_logger.info(f"Timeout exception on {equipment['secret']} ({equipment['ip']})")

    except AuthenticationException:
        main_logger.info(f"Authentication failed on {equipment['secret']} ({equipment['ip']})")

    except SSHException:
        main_logger.info(f"Error reading SSH protocol banner on {equipment['secret']} ({equipment['ip']})")

    except:
        main_logger.info(f"An error has occurred on {equipment['secret']} ({equipment['ip']})")

    else:
        hostname, model, serial_number, ios = equip_info(output, hostname)

        write_output(equipment, model, serial_number, ios, hostname)



def equip_info(output, hostname):
    re_model = r'(Model number.*)'
    re_model_res = re.search(re_model, output)
    re_model_res = re_model_res.group()
    model = re_model_res.partition(":")[2].strip()

    re_serial_number = r'(System serial number.*)'
    re_serial_number_res = re.search(re_serial_number, output)
    re_serial_number_res = re_serial_number_res.group()
    serial_number = re_serial_number_res.partition(":")[2].strip()

    re_ios = r'(System image file is.*)'
    re_ios_res = re.search(re_ios, output)
    re_ios_res = re_ios_res.group()
    pos_ios = re_ios_res.partition(":")[2].find("/", 1)
    ios = re_ios_res.partition(":")[2][pos_ios+1:-1].replace("/", "").strip()

    hostname = hostname[:-1].strip()

    return hostname,model,serial_number,ios



def write_output(equipment, model, serial_number, ios, hostname):
    with open(os.path.join('Outputs', 'output-SW-info.txt'), 'at') as f:
        f.write(f"{equipment['secret']} ({equipment['ip']}): {model}\t{serial_number}\t{ios}\t{hostname}\n")
        main_logger.info(f"Completed: {equipment['secret']} ({equipment['ip']})")



if __name__ == '__main__':
    cls()
    time_start = datetime.datetime.now()
    time_start = time_start.strftime('%H:%M:%S %d/%m/%Y')
    perf_counter_start = time.perf_counter()

    num_equips = env_exec()

    perf_counter_stop = time.perf_counter()
    time_stop = datetime.datetime.now()
    time_stop = time_stop.strftime('%H:%M:%S %d/%m/%Y')

    main_logger.info(f'Number of equipments: {num_equips}')
    main_logger.info(f'Finished in {(perf_counter_stop - perf_counter_start):.3f} second(s)')
    main_logger.info(f'Started at {time_start} and finished at {time_stop}\n')
