import os
import subprocess
import sys

import openpyxl


def main():
    if sys.platform in ("linux", "darwin"):
        excel = "/mnt/c/Users/joao.caldeira.ext/Portugal Telecom/SPMS RIS2020 - NOC - General/Gestao_COS_RIS.xlsm"
    elif sys.platform == "win32":
        excel = r"C:\Users\joao.caldeira.ext\Portugal Telecom\SPMS RIS2020 - NOC - General\Gestao_COS_RIS.xlsm"

    excelFileName = os.path.basename(excel)
    pathToExcelFile = os.path.dirname(excel)

    chars_a_remover = ['á','Á','à','À','ã','Ã','â','Â','é','É','è','È','ê','Ê','í','Í','ó','Ó','õ','Õ','ô','Ô','ú','Ú','û','Û','ù','Ù','ç','s/n',' ']
    chars_a_inserir = ['a','A','a','A','a','A','a','A','e','E','e','E','e','E','i','I','o','O','o','O','o','O','u','U','u','U','u','U','c','','_']

    try:
        wb = openpyxl.load_workbook(excel, data_only=True, read_only=True)
        delTempFileFlag = False
    except PermissionError as e:
        print(f"File is open, so cannot be accessed:\n{e.filename}\n")
        createTempFile(pathToExcelFile, os.getcwd(), excelFileName)
        print("Temporary copy of file created to work with\n")
        wb = openpyxl.load_workbook(excelFileName, data_only=True, read_only=True)
        delTempFileFlag = True

    except FileNotFoundError as e:
        sys.exit(f"File not found: {e.filename}")


    sheet = wb["Sites RIS2020"]

    while True:
        user_input = input('Site: ')
        if user_input == '':
            continue
        if user_input.lower() == 'exit':
            break
        for row in sheet.iter_rows(min_row=3, values_only=True):
            if row[0] == user_input:
                id_site = row[0]
                nome_local = site_info(chars_a_remover, chars_a_inserir, row[1])
                ip = str(row[18]).replace('/32', '')
                ars = row[44]

                print(f'{id_site}-{nome_local}\t{ip}\t{ars}\n')



    wb.close()
    if delTempFileFlag == True:
        os.remove(excelFileName)


def site_info(chars_a_remover, chars_a_inserir, nome_local):
    for i in range(0,len(chars_a_remover)):
        nome_local = nome_local.replace(chars_a_remover[i],chars_a_inserir[i]).strip()
    return nome_local


def cls():
    if sys.platform in ("linux", "darwin"):
        subprocess.run(["clear"])
    elif sys.platform == "win32":
        subprocess.run(["cls"], shell=True)



def createTempFile(pathToExcelFile, cwd, excelFileName):
    if sys.platform in ("linux", "darwin"):
        subprocess.run(["cp", pathToExcelFile, cwd, excelFileName])
    elif sys.platform == "win32":
        subprocess.run(["robocopy", pathToExcelFile, cwd, excelFileName], shell=True, stdout=subprocess.DEVNULL)



if __name__ == "__main__":
    cls()
    main()
