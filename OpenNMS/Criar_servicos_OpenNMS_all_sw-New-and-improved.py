"""
https://stackoverflow.com/questions/56770093/read-from-excel-file-that-is-open-in-python
"""

import datetime
import ipaddress
import os
import subprocess
import sys

import openpyxl


def main():
    if sys.platform in ("linux", "darwin"):
        excel = "/mnt/c/Users/joao.caldeira.ext/SPMS - Serviços Partilhados do Ministério da Saúde, EPE/SPMS DSI RDIS - RIS Corporate - RIS2020 Cadastro/RIS2020 - Cadastro.xlsx"
    elif sys.platform == "win32":
        excel = r"C:\Users\joao.caldeira.ext\SPMS - Serviços Partilhados do Ministério da Saúde, EPE\SPMS DSI RDIS - RIS Corporate - RIS2020 Cadastro\RIS2020 - Cadastro.xlsx"

    excelFileName = os.path.basename(excel)
    pathToExcelFile = os.path.dirname(excel)

    chars = {
        'á': 'a',
        'Á': 'A',
        'à': 'a',
        'À': 'A',
        'ã': 'a',
        'Ã': 'A',
        'â': 'a',
        'Â': 'A',
        'é': 'e',
        'É': 'E',
        'è': 'e',
        'È': 'E',
        'ê': 'e',
        'Ê': 'E',
        'í': 'i',
        'Í': 'I',
        'ó': 'o',
        'Ó': 'O',
        'õ': 'o',
        'Õ': 'O',
        'ô': 'o',
        'Ô': 'O',
        'ú': 'u',
        'Ú': 'U',
        'û': 'u',
        'Û': 'U',
        'ù': 'u',
        'Ù': 'U',
        'ç': 'c',
        's/n': '',
        'S/N': '',
        'º': 'r',
        '\'': ''
    }

    try:
        wb = openpyxl.load_workbook(excel, data_only=True, read_only=True)
        delTempFileFlag = False
    except PermissionError as e:
        print(f"File is in use, so cannot be accessed:\n{e.filename}\n")
        createTempFile(pathToExcelFile, os.getcwd(), excelFileName)
        print("Temporary copy of file created to work with\n")
        wb = openpyxl.load_workbook(excelFileName, data_only=True, read_only=True)
        delTempFileFlag = True

    except FileNotFoundError as e:
        sys.exit(f"File not found: {e.filename}")


    sheet = wb["RIS2020"]
    counter = 0
    excluded_ids = ['1626']
    included_ids = []
    excluded_ips = []
    included_ips = ['10.33.7.241']
    with open(os.path.join('Outputs', 'output-servicos.txt'), 'wt') as f:
        for row in sheet.iter_rows(min_row=3, values_only=True):
            if type(row[26]) == datetime.datetime:
            # if row[26] is not None:
                data_presi = row[26].date()
                ip = str(row[24])
                hostname = str(row[23])
                site_id = row[0]
                # if validate_ip(ip) and validate_date_inter(data_presi) and site_id not in excluded_ids and ip not in excluded_ips:
                # if validate_ip(ip) and hostname == '0073-B01-SW01':
                # if validate_ip(ip) and site_id == '2605':
                if validate_ip(ip) and (site_id in included_ids or ip in included_ips):
                    # if str(hostname[:4]) == site_id:
                    entidade, morada, cp, localidade, latitude, longitude, requisition, modelo, serial, node_id = site_info(row, chars)

                    f.write(f"/opt/opennms/bin/provision.pl service add '{requisition}' {node_id} {ip} ICMP\n")
                    f.write(f"/opt/opennms/bin/provision.pl service add '{requisition}' {node_id} {ip} SNMP\n")
                    f.write(f"/opt/opennms/bin/provision.pl category remove '{requisition}' {node_id} '{entidade}'\n")
                    f.write(f"/opt/opennms/bin/provision.pl category add '{requisition}' {node_id} '{entidade.upper()}'\n")
                    f.write(f"/opt/opennms/bin/provision.pl category add '{requisition}' {node_id} 'Production'\n")
                    f.write(f"/opt/opennms/bin/provision.pl category add '{requisition}' {node_id} 'Switches'\n")
                    f.write(f"/opt/opennms/bin/provision.pl asset add '{requisition}' {node_id} address1 '{morada}'\n")
                    f.write(f"/opt/opennms/bin/provision.pl asset add '{requisition}' {node_id} zip '{cp}'\n")
                    f.write(f"/opt/opennms/bin/provision.pl asset add '{requisition}' {node_id} city '{localidade}'\n")
                    f.write(f"/opt/opennms/bin/provision.pl asset add '{requisition}' {node_id} modelNumber '{modelo}'\n")
                    f.write(f"/opt/opennms/bin/provision.pl asset add '{requisition}' {node_id} serialNumber '{serial}'\n")
                    f.write(f"/opt/opennms/bin/provision.pl asset add '{requisition}' {node_id} latitude '{latitude}'\n")
                    f.write(f"/opt/opennms/bin/provision.pl asset add '{requisition}' {node_id} longitude '{longitude}'\n")
                    f.write(f"/opt/opennms/bin/provision.pl asset add '{requisition}' {node_id} country Portugal\n")
                    f.write(f"/opt/opennms/bin/provision.pl requisition import {requisition}\n\n")

                    print(f"Completed: {site_id} ({ip})")

                    counter += 1


    wb.close()
    if delTempFileFlag == True:
        os.remove(excelFileName)

    print(f'Generated {counter} configurations')


def site_info(row, chars):
    entidade = str(row[1]).replace(" ","_").strip()
    morada = str(row[4]).strip()
    cp = str(row[5]).strip()
    localidade = str(row[6]).strip()
    latitude = str(row[7]).strip()
    longitude = str(row[8]).strip()
    requisition = str(row[28]).strip()
    modelo = str(row[20]).strip()
    serial = str(row[21]).strip()
    node_id = str(row[27]).strip()
    for char in chars:
        if char in chars:
            morada = morada.replace(char, chars[char]).strip()
    for char in chars:
        if char in chars:
            localidade = localidade.replace(char, chars[char]).strip()
    return entidade, morada, cp, localidade, latitude, longitude, requisition, modelo, serial, node_id



def validate_date_inter(date):
    low_date_lim = datetime.date(2021, 6, 21)
    high_date_lim = datetime.date(2021, 6, 22)
    return low_date_lim <= date <= high_date_lim



def validate_ip(ip):
    try:
        ipaddress.ip_address(ip)
    except:
        return False
    else:
        return True



def cls():
    if sys.platform in ("linux", "darwin"):
        subprocess.run(["clear"])
    elif sys.platform == "win32":
        subprocess.run(["cls"], shell=True)



def createTempFile(pathToExcelFile, cwd, excelFileName):
    if sys.platform in ("linux", "darwin"):
        sys.exit(f"Copy mechanism of in use files not implemented in non Windows platforms")
    elif sys.platform == "win32":
        subprocess.run(["robocopy", pathToExcelFile, cwd, excelFileName], shell=True, stdout=subprocess.DEVNULL)



if __name__ == "__main__":
    cls()
    main()
