"""
https://stackoverflow.com/questions/56770093/read-from-excel-file-that-is-open-in-python
"""

import datetime
import ipaddress
import logging
import os
import subprocess
import sys

import openpyxl

main_logger = logging.getLogger(__name__)
main_logger.setLevel(logging.DEBUG)

file_formatter = logging.Formatter('%(asctime)s||%(levelname)s||%(name)s||%(message)s')
stream_formatter = logging.Formatter('%(message)s')

file_handler = logging.FileHandler(f'{os.path.join("Logs",os.path.splitext(os.path.basename(__file__))[0])}.log')
file_handler.setLevel(logging.DEBUG)
file_handler.setFormatter(file_formatter)

stream_handler = logging.StreamHandler()
stream_handler.setLevel(logging.INFO)
stream_handler.setFormatter(stream_formatter)

main_logger.addHandler(file_handler)
main_logger.addHandler(stream_handler)



def main():
    if sys.platform in ("linux", "darwin"):
        excel = "/mnt/c/Users/joao.caldeira.ext/SPMS - Serviços Partilhados do Ministério da Saúde, EPE/SPMS DSI RDIS - RIS Corporate - RIS2020 Cadastro/RIS2020 - Cadastro.xlsx"
        excel2 = "/mnt/c/Users/joao.caldeira.ext/SPMS - Serviços Partilhados do Ministério da Saúde, EPE/SPMS DSI RDIS - RIS Corporate - RIS2020 Cadastro/Opennms/raw_data.xlsx"

    elif sys.platform == "win32":
        excel = r"C:\Users\joao.caldeira.ext\SPMS - Serviços Partilhados do Ministério da Saúde, EPE\SPMS DSI RDIS - RIS Corporate - RIS2020 Cadastro\RIS2020 - Cadastro.xlsx"
        excel2 = r"C:\Users\joao.caldeira.ext\SPMS - Serviços Partilhados do Ministério da Saúde, EPE\SPMS DSI RDIS - RIS Corporate - RIS2020 Cadastro\Opennms\raw_data.xlsx"

    chars = {
        'á': 'a',
        'Á': 'A',
        'à': 'a',
        'À': 'A',
        'ã': 'a',
        'Ã': 'A',
        'â': 'a',
        'Â': 'A',
        'é': 'e',
        'É': 'E',
        'è': 'e',
        'È': 'E',
        'ê': 'e',
        'Ê': 'E',
        'í': 'i',
        'Í': 'I',
        'ó': 'o',
        'Ó': 'O',
        'õ': 'o',
        'Õ': 'O',
        'ô': 'o',
        'Ô': 'O',
        'ú': 'u',
        'Ú': 'U',
        'û': 'u',
        'Û': 'U',
        'ù': 'u',
        'Ù': 'U',
        'ç': 'c',
        's/n': '',
        'S/N': '',
        'º': '',
        'ª': '',
        '\'': '',
        '/': ' ',
        '\\': ' '
    }

    excelFileName = os.path.basename(excel)
    pathToExcelFile = os.path.dirname(excel)

    excelFileName2 = os.path.basename(excel2)
    pathToExcelFile2 = os.path.dirname(excel2)

    # Open Excel
    try:
        wb = openpyxl.load_workbook(excel, data_only=True)
        delTempFileFlag = False
    except PermissionError as e:
        main_logger.info(f'File is in use, so cannot be accessed:\n{e.filename}\n')
        createTempFile(pathToExcelFile, os.getcwd(), excelFileName)
        main_logger.info('Temporary copy of the file created to work with\n')
        wb = openpyxl.load_workbook(excelFileName, data_only=True)
        delTempFileFlag = True
    except FileNotFoundError as e:
        sys.exit(f'File not found: {e.filename}')


    # Open Excel 2
    try:
        wb2 = openpyxl.load_workbook(excel2, data_only=True)
        delTempFileFlag2 = False
    except PermissionError as e:
        main_logger.info(f'File is in use, so cannot be accessed:\n{e.filename}\n')
        createTempFile(pathToExcelFile2, os.getcwd(), excelFileName2)
        main_logger.info('Temporary copy of the file created to work with\n')
        wb2 = openpyxl.load_workbook(excelFileName2, data_only=True)
        delTempFileFlag2 = True
    except FileNotFoundError as e:
        sys.exit(f'File not found: {e.filename}')


    sheet = wb["RIS2020"]
    sheet2 = wb2["Sheet1"]
    counter = 0
    included_ids = []
    included_node_ids = []
    with open(os.path.join('Outputs', 'output-UPS.txt'), 'wt') as f:
        site_id = ''
        for row2 in sheet2.iter_rows(min_row=2, values_only=True):
            # node_id = str(row2[1]).zfill(4)
            node_id = str(row2[1])
            node_label = str(row2[0])
            for row in sheet.iter_rows(min_row=3, values_only=True):
                if node_id in included_node_ids:
                    if str(row[0]) == site_id:
                        continue
                    if validate_ip(node_label):
                        ip_lan_raw = node_label[:node_label.rfind('.')]
                        ip_lan_cadastro = str(row[24])[:str(row[24]).rfind('.')]
                    if ip_lan_raw == ip_lan_cadastro:
                        site_id = str(row[0])
                        node_label = f'{site_id}-UPS'
                    else:
                        continue

                    entidade, requisition = site_info(row, row2)
                    writeoutput(f, node_id, node_label, ip_lan_cadastro, entidade, requisition)
                    main_logger.info(f"Completed: {node_id} ({node_label})")

                    counter += 1

                # elif (node_label[:4] in included_ids) and ('UPS' in node_label):
                elif 'UPS' in node_label:
                    # for row in sheet.iter_rows(min_row=3, values_only=True):
                    if str(row[0]) == site_id:
                        continue
                    site_id = str(row[0])
                    ip_lan = str(row[24]).replace('/24','').strip()
                    entidade, requisition = site_info(row, row2)
                    writeoutput(f, node_id, node_label, ip_lan, entidade, requisition)
                    main_logger.info(f"Completed: {node_id} ({node_label})")




    wb.close()
    if delTempFileFlag == True:
        os.remove(excelFileName)

    print(f'\nGenerated {counter} configurations')


def writeoutput(f, node_id, node_label, ip, entidade, requisition):
    f.write(f"/opt/opennms/bin/provision.pl node set '{requisition}' {node_id} node-label '{node_label}'\n")
    f.write(f"/opt/opennms/bin/provision.pl service add '{requisition}' {node_id} {ip} ICMP\n")
    f.write(f"/opt/opennms/bin/provision.pl service add '{requisition}' {node_id} {ip} SNMP\n")
    f.write(f"/opt/opennms/bin/provision.pl service add '{requisition}' {node_id} {ip} HTTP\n")
    f.write(f"/opt/opennms/bin/provision.pl service remove '{requisition}' {node_id} {ip} SSH\n")
    f.write(f"/opt/opennms/bin/provision.pl category add '{requisition}' {node_id} '{entidade.upper()}'\n")
    f.write(f"/opt/opennms/bin/provision.pl category add '{requisition}' {node_id} 'Production'\n")
    f.write(f"/opt/opennms/bin/provision.pl category add '{requisition}' {node_id} 'UPS'\n")
    f.write(f"/opt/opennms/bin/provision.pl requisition import {requisition}\n\n")


def site_info(row, row2):
    entidade = str(row[1]).replace(" ","_").strip()
    requisition = str(row2[22]).strip()
    return entidade, requisition



def validate_date_inter(date):
    low_date_lim = datetime.date(2021, 5, 24)
    high_date_lim = datetime.date(2021, 5, 28)
    return low_date_lim <= date <= high_date_lim



def validate_ip(ip):
    try:
        ipaddress.ip_address(ip)
    except:
        return False
    else:
        return True



def cls():
    if sys.platform in ("linux", "darwin"):
        subprocess.run(["clear"])
    elif sys.platform == "win32":
        subprocess.run(["cls"], shell=True)



def createTempFile(pathToExcelFile, cwd, excelFileName):
    if sys.platform in ("linux", "darwin"):
        sys.exit(f"Copy mechanism of in use files not implemented in non Windows platforms")
    elif sys.platform == "win32":
        subprocess.run(["robocopy", pathToExcelFile, cwd, excelFileName], shell=True, stdout=subprocess.DEVNULL)



if __name__ == "__main__":
    cls()
    main()
