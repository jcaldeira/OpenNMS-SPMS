import ipaddress
import sys

import openpyxl


def main():
    excel = "C:\\Users\\joao.caldeira.ext\\SPMS - Serviços Partilhados do Ministério da Saúde, EPE\\SPMS DSI RDIS - RIS Corporate - RIS2020 Cadastro\\RIS2020 - Cadastro.xlsx"
    chars_a_remover = ['á','à','ã','â','é','è','ê','í','ó','õ','ô','ú','û','ù','ç','s/n']
    chars_a_inserir = ['a','a','a','a','e','e','e','i','o','o','o','u','u','u','c','']

    try:
        wb = openpyxl.load_workbook(excel, data_only=True)
    except Exception as e:
        sys.exit(f"{e}\nImpossible to open excel file: {excel}")
    else:
        sheet = wb["RIS2020"]

    for row in sheet.iter_rows(min_row=3, values_only=True):
        ip = str(row[24])
        hostname = str(row[23])
        site_id = hostname[0:4]
        if validate_ip(ip) and hostname == '2014-B01-SW01':
            if 'B01-SW01' not in hostname:
                for row2 in sheet.iter_rows(min_row=3, values_only=True):
                    if str(row[0]) == site_id:
                        entidade, morada, cp, localidade, latitude, longitude = site_info(chars_a_remover, chars_a_inserir, row)

            else:
                entidade, morada, cp, localidade, latitude, longitude = site_info(chars_a_remover, chars_a_inserir, row)

            aces = str(row[28])
            modelo = str(row[20])
            serial = str(row[21])
            node_id = str(row[27])

            print(f"/opt/opennms/bin/provision.pl service add '{aces}' {node_id} {ip} ICMP")
            print(f"/opt/opennms/bin/provision.pl service add '{aces}' {node_id} {ip} SNMP")
            print(f"/opt/opennms/bin/provision.pl category remove '{aces}' {node_id} '{entidade}'")
            print(f"/opt/opennms/bin/provision.pl category add '{aces}' {node_id} '{entidade.upper()}'")
            print(f"/opt/opennms/bin/provision.pl category add '{aces}' {node_id} 'Production'")
            print(f"/opt/opennms/bin/provision.pl category add '{aces}' {node_id} 'Switches'")
            print(f"/opt/opennms/bin/provision.pl service add '{aces}' {node_id} address1 '{morada}'")
            print(f"/opt/opennms/bin/provision.pl service add '{aces}' {node_id} zip '{cp}'")
            print(f"/opt/opennms/bin/provision.pl service add '{aces}' {node_id} city '{localidade}'")
            print(f"/opt/opennms/bin/provision.pl service add '{aces}' {node_id} modelNumber '{modelo}'")
            print(f"/opt/opennms/bin/provision.pl service add '{aces}' {node_id} serialNumber '{serial}'")
            print(f"/opt/opennms/bin/provision.pl service add '{aces}' {node_id} latitude '{latitude}'")
            print(f"/opt/opennms/bin/provision.pl service add '{aces}' {node_id} longitude '{longitude}'")
            print(f"/opt/opennms/bin/provision.pl service add '{aces}' {node_id} country Portugal")
            print(f"/opt/opennms/bin/provision.pl requisition import {aces}\n")


def site_info(chars_a_remover, chars_a_inserir, row):
    entidade = str(row[1]).replace(" ","_")
    morada = str(row[4])
    cp = str(row[5])
    localidade = str(row[6])
    latitude = str(row[7])
    longitude = str(row[8])
    local = str(row[3])
    for i in range(0,len(chars_a_remover)):
        morada = morada.replace(chars_a_remover[i],chars_a_inserir[i])
        localidade = localidade.replace(chars_a_remover[i], chars_a_inserir[i])
    return entidade, morada, cp, localidade, latitude, longitude


def validate_ip(ip):
    try:
        ipaddress.ip_address(ip)
    except:
        return False
    else:
        return True



if __name__ == "__main__":
    main()
